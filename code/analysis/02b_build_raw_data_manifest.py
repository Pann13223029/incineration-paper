"""Build an auditable provenance manifest for the raw MOE workbooks.

The downloader does not persist HTTP response metadata or retrieval times. This
stage therefore distinguishes configured source information from facts that can
be recovered from the files now present in ``data/raw``. Missing provenance is
recorded explicitly rather than inferred.
"""

from __future__ import annotations

from collections import Counter
from collections.abc import Mapping
import hashlib
import importlib.util
import json
import subprocess
import sys
from pathlib import Path
from types import ModuleType
from typing import Any

import pandas as pd
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
ANALYSIS_DIR = Path(__file__).resolve().parent
RAW_DIR = REPO_ROOT / "data" / "raw" / "facility_annual"
OUTPUT_DIR = REPO_ROOT / "output"
WORKBOOK_MANIFEST_PATH = OUTPUT_DIR / "raw_data_manifest.csv"
SCHEMA_MAP_PATH = OUTPUT_DIR / "raw_workbook_schema_map.csv"
REPORT_PATH = OUTPUT_DIR / "raw_data_provenance.md"
CANONICAL_FISCAL_YEARS = tuple(range(2005, 2025))
EXPECTED_WORKBOOK_COUNT = 20
WORKBOOK_EXTENSIONS = ("xlsx", "xls")

if len(CANONICAL_FISCAL_YEARS) != EXPECTED_WORKBOOK_COUNT:
    raise RuntimeError("Canonical ingestion window must contain exactly 20 fiscal years")

if str(ANALYSIS_DIR) not in sys.path:
    sys.path.insert(0, str(ANALYSIS_DIR))

from panel_utils import write_stage_manifest  # noqa: E402


class IngestionPreflightError(ValueError):
    """Raised when canonical raw inputs cannot support a complete rebuild."""


def load_analysis_module(module_name: str, filename: str) -> ModuleType:
    """Load a numbered analysis module without relying on an importable name."""
    path = ANALYSIS_DIR / filename
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load analysis configuration from {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def validate_ingestion_configuration(
    downloader: ModuleType,
    parser: ModuleType,
) -> list[tuple[int, str]]:
    """Validate the downloader and parser against the fixed canonical contract."""
    configured = getattr(downloader, "YEARS", None)
    if not isinstance(configured, Mapping):
        raise IngestionPreflightError(
            "Downloader YEARS configuration must be a mapping of fiscal year to era code"
        )

    configured_years = list(configured.keys())
    invalid_years = [year for year in configured_years if type(year) is not int]
    if invalid_years:
        raise IngestionPreflightError(
            "Downloader YEARS contains non-integer fiscal years: "
            + ", ".join(sorted(map(repr, invalid_years)))
        )

    duplicate_years = sorted(
        year for year, count in Counter(configured_years).items() if count > 1
    )
    expected_years = set(CANONICAL_FISCAL_YEARS)
    actual_years = set(configured_years)
    missing_years = sorted(expected_years - actual_years)
    unexpected_years = sorted(actual_years - expected_years)
    if (
        duplicate_years
        or missing_years
        or unexpected_years
        or len(configured_years) != EXPECTED_WORKBOOK_COUNT
    ):
        raise IngestionPreflightError(
            "Downloader fiscal-year configuration is not canonical: "
            f"configured_count={len(configured_years)}, "
            f"expected_count={EXPECTED_WORKBOOK_COUNT}, "
            f"duplicates={duplicate_years or 'none'}, "
            f"missing={missing_years or 'none'}, "
            f"unexpected={unexpected_years or 'none'}"
        )

    configured_items = sorted(configured.items())
    invalid_era_codes = [
        fiscal_year
        for fiscal_year, era_code in configured_items
        if not isinstance(era_code, str) or not era_code.strip()
    ]
    if invalid_era_codes:
        raise IngestionPreflightError(
            "Downloader YEARS has blank/non-string era codes for: "
            + ", ".join(f"FY{year}" for year in invalid_era_codes)
        )
    duplicate_era_codes = sorted(
        era_code
        for era_code, count in Counter(
            era_code for _, era_code in configured_items
        ).items()
        if count > 1
    )
    if duplicate_era_codes:
        raise IngestionPreflightError(
            "Downloader YEARS maps multiple fiscal years to the same era code: "
            + ", ".join(duplicate_era_codes)
        )

    parser_years = tuple(getattr(parser, "CANONICAL_FISCAL_YEARS", ()))
    if parser_years != CANONICAL_FISCAL_YEARS:
        raise IngestionPreflightError(
            "Parser canonical fiscal years disagree with provenance preflight: "
            f"parser={list(parser_years)}, expected={list(CANONICAL_FISCAL_YEARS)}"
        )

    configured_fields = [field_name for field_name, _ in parser.COLUMN_DEFS]
    duplicate_fields = sorted(
        field_name
        for field_name, count in Counter(configured_fields).items()
        if count > 1
    )
    required_fields = tuple(getattr(parser, "REQUIRED_STANDARDIZED_FIELDS", ()))
    optional_fields = tuple(getattr(parser, "OPTIONAL_STANDARDIZED_FIELDS", ()))
    optional_start_years = getattr(parser, "OPTIONAL_FIELD_FIRST_FISCAL_YEAR", {})
    if duplicate_fields:
        raise IngestionPreflightError(
            "Parser COLUMN_DEFS contains duplicate standardized fields: "
            + ", ".join(duplicate_fields)
        )
    if not required_fields:
        raise IngestionPreflightError(
            "Parser must declare nonempty REQUIRED_STANDARDIZED_FIELDS"
        )
    if set(required_fields) & set(optional_fields):
        overlap = sorted(set(required_fields) & set(optional_fields))
        raise IngestionPreflightError(
            "Parser required and optional standardized fields overlap: "
            + ", ".join(overlap)
        )
    if set(optional_start_years) != set(optional_fields):
        raise IngestionPreflightError(
            "Parser optional-field start-year contract is incomplete: "
            f"configured={sorted(optional_start_years)}, "
            f"expected={sorted(optional_fields)}"
        )
    invalid_optional_start_years = {
        field_name: first_year
        for field_name, first_year in optional_start_years.items()
        if type(first_year) is not int or first_year not in CANONICAL_FISCAL_YEARS
    }
    if invalid_optional_start_years:
        raise IngestionPreflightError(
            "Parser optional-field start years are outside the canonical window: "
            f"{invalid_optional_start_years}"
        )
    if set(required_fields) | set(optional_fields) != set(configured_fields):
        unclassified = sorted(
            set(configured_fields) - set(required_fields) - set(optional_fields)
        )
        unknown = sorted(
            (set(required_fields) | set(optional_fields)) - set(configured_fields)
        )
        raise IngestionPreflightError(
            "Parser standardized-field contract does not partition COLUMN_DEFS: "
            f"unclassified={unclassified or 'none'}, unknown={unknown or 'none'}"
        )

    return [(int(year), str(era_code)) for year, era_code in configured_items]


def preflight_workbook_inventory(
    downloader: ModuleType,
    parser: ModuleType,
    raw_dir: Path = RAW_DIR,
) -> list[tuple[int, str, Path]]:
    """Require one present, nonempty workbook for every canonical fiscal year."""
    configured_items = validate_ingestion_configuration(downloader, parser)
    if not raw_dir.is_dir():
        raise IngestionPreflightError(
            f"Canonical raw workbook directory does not exist: {raw_dir}"
        )

    inventory: list[tuple[int, str, Path]] = []
    missing_years: list[int] = []
    duplicate_files: dict[int, list[str]] = {}
    empty_files: list[str] = []
    non_file_candidates: list[str] = []
    configured_candidates: set[Path] = set()

    for fiscal_year, era_code in configured_items:
        candidates = [
            raw_dir / f"fy{fiscal_year}_incineration.{extension}"
            for extension in WORKBOOK_EXTENSIONS
        ]
        configured_candidates.update(path.resolve() for path in candidates)
        non_file_candidates.extend(
            path.name for path in candidates if path.exists() and not path.is_file()
        )
        existing = [path for path in candidates if path.is_file()]
        if not existing:
            missing_years.append(fiscal_year)
            continue
        if len(existing) > 1:
            duplicate_files[fiscal_year] = sorted(path.name for path in existing)
            continue

        path = existing[0]
        if path.stat().st_size <= 0:
            empty_files.append(path.name)
            continue
        inventory.append((fiscal_year, era_code, path))

    unexpected_files = sorted(
        path.name
        for path in raw_dir.iterdir()
        if path.is_file() and path.resolve() not in configured_candidates
    )
    problems: list[str] = []
    if missing_years:
        problems.append(
            "missing=" + ",".join(f"FY{year}" for year in missing_years)
        )
    if duplicate_files:
        duplicate_text = ";".join(
            f"FY{year}:[{','.join(names)}]"
            for year, names in sorted(duplicate_files.items())
        )
        problems.append(f"duplicates={duplicate_text}")
    if empty_files:
        problems.append("empty=" + ",".join(sorted(empty_files)))
    if non_file_candidates:
        problems.append(
            "not_regular_files=" + ",".join(sorted(non_file_candidates))
        )
    if unexpected_files:
        problems.append("unexpected=" + ",".join(unexpected_files))
    if len(inventory) != EXPECTED_WORKBOOK_COUNT:
        problems.append(
            f"valid_workbook_count={len(inventory)},expected={EXPECTED_WORKBOOK_COUNT}"
        )
    if problems:
        raise IngestionPreflightError(
            "Canonical raw-workbook inventory preflight failed: " + " | ".join(problems)
        )

    inventory_years = [fiscal_year for fiscal_year, _, _ in inventory]
    duplicate_inventory_years = sorted(
        year for year, count in Counter(inventory_years).items() if count > 1
    )
    if inventory_years != list(CANONICAL_FISCAL_YEARS):
        raise IngestionPreflightError(
            "Canonical raw-workbook inventory has invalid fiscal-year ordering/coverage: "
            f"years={inventory_years}, duplicates={duplicate_inventory_years or 'none'}"
        )
    return inventory


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def repository_commit_timestamp(path: Path) -> str:
    """Return deterministic repository history time rather than checkout mtime."""
    result = subprocess.run(
        [
            "git",
            "log",
            "-1",
            "--format=%cI",
            "--",
            path.relative_to(REPO_ROOT).as_posix(),
        ],
        cwd=REPO_ROOT,
        check=False,
        capture_output=True,
        text=True,
    )
    value = result.stdout.strip()
    return value if result.returncode == 0 and value else "unavailable"


def detect_header_match(
    frame: pd.DataFrame,
    keywords: list[str],
    normalize: Any,
    max_rows: int = 6,
) -> dict[str, Any] | None:
    """Return the exact first header match used by the parser search order."""
    for row_idx in range(min(max_rows, frame.shape[0])):
        for col_idx in range(frame.shape[1]):
            value = normalize(frame.iloc[row_idx, col_idx])
            for keyword in keywords:
                if keyword in value:
                    return {
                        "header_row_zero_based": row_idx,
                        "header_row_excel": row_idx + 1,
                        "column_index_zero_based": col_idx,
                        "column_excel": get_column_letter(col_idx + 1),
                        "matched_keyword": keyword,
                        "matched_header_text": value,
                    }
    return None


def unavailable_mapping_row(fiscal_year: int, field_name: str) -> dict[str, Any]:
    return {
        "fiscal_year": fiscal_year,
        "standardized_field": field_name,
        "mapping_status": "unavailable_not_detected",
        "header_row_zero_based": "unavailable",
        "header_row_excel": "unavailable",
        "column_index_zero_based": "unavailable",
        "column_excel": "unavailable",
        "parser_selected_column_index_zero_based": "unavailable",
        "parser_selected_column_excel": "unavailable",
        "parser_selected_header_text": "unavailable",
        "parser_column_adjustment": "unavailable",
        "matched_keyword": "unavailable",
        "matched_header_text": "unavailable",
    }


def inspect_workbook(
    fiscal_year: int,
    path: Path,
    era_code: str,
    downloader: ModuleType,
    parser: ModuleType,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    stat = path.stat()
    if stat.st_size <= 0:
        raise IngestionPreflightError(
            f"FY{fiscal_year} workbook is zero bytes: {path}"
        )

    engine = "openpyxl" if path.suffix.lower() == ".xlsx" else "xlrd"
    with pd.ExcelFile(path, engine=engine) as workbook:
        sheet_names = list(workbook.sheet_names)
    if not sheet_names:
        raise IngestionPreflightError(
            f"FY{fiscal_year} workbook contains no worksheets: {path}"
        )

    raw = pd.read_excel(path, sheet_name=0, header=None, engine=engine)
    if raw.empty or raw.shape[1] == 0:
        raise IngestionPreflightError(
            f"FY{fiscal_year} workbook first sheet is empty: {path}"
        )
    data_start = int(parser.find_data_start(raw))
    if data_start < 0 or data_start >= len(raw):
        raise IngestionPreflightError(
            f"FY{fiscal_year} workbook has no candidate rows at detected data start "
            f"{data_start} (zero-based): {path}"
        )
    candidate_data_rows = int(len(raw) - data_start)

    parser_columns = {
        field_name: parser.find_column(raw, list(keywords))
        for field_name, keywords in parser.COLUMN_DEFS
    }
    final_parser_columns = dict(parser_columns)
    if parser_columns.get("prefecture") is not None:
        initial_column = int(parser_columns["prefecture"])
        data = raw.iloc[data_start:].reset_index(drop=True)
        sample = parser.normalize(data.iloc[0, initial_column])
        adjacent_column = initial_column + 1
        if sample and sample.isdigit() and adjacent_column < data.shape[1]:
            adjacent_sample = parser.normalize(data.iloc[0, adjacent_column])
            if adjacent_sample and not adjacent_sample.isdigit():
                final_parser_columns["prefecture"] = adjacent_column

    mapping_rows: list[dict[str, Any]] = []
    mapping_for_hash: dict[str, Any] = {}
    for field_name, keywords in parser.COLUMN_DEFS:
        match = detect_header_match(raw, list(keywords), parser.normalize)
        parser_column = parser_columns[field_name]
        selected_column = final_parser_columns[field_name]
        if match is None:
            row = unavailable_mapping_row(fiscal_year, field_name)
            if parser_column is not None:
                raise ValueError(
                    f"Mapping audit disagrees with parser for FY{fiscal_year} {field_name}"
                )
            mapping_for_hash[field_name] = None
        else:
            if int(match["column_index_zero_based"]) != int(parser_column):
                raise ValueError(
                    f"Mapping audit disagrees with parser for FY{fiscal_year} {field_name}"
                )
            row = {
                "fiscal_year": fiscal_year,
                "standardized_field": field_name,
                "mapping_status": "detected_by_parser_keyword_search",
                **match,
                "parser_selected_column_index_zero_based": int(selected_column),
                "parser_selected_column_excel": get_column_letter(
                    int(selected_column) + 1
                ),
                "parser_selected_header_text": parser.normalize(
                    raw.iloc[int(match["header_row_zero_based"]), int(selected_column)]
                ),
                "parser_column_adjustment": (
                    "prefecture_code_to_adjacent_name_column"
                    if int(selected_column) != int(parser_column)
                    else "none"
                ),
            }
            mapping_for_hash[field_name] = {
                "row": match["header_row_zero_based"],
                "detected_column": match["column_index_zero_based"],
                "selected_column": int(selected_column),
                "selected_header": parser.normalize(
                    raw.iloc[int(match["header_row_zero_based"]), int(selected_column)]
                ),
                "keyword": match["matched_keyword"],
                "header": match["matched_header_text"],
            }
        mapping_rows.append(row)

    detected = sum(row["mapping_status"].startswith("detected") for row in mapping_rows)
    unavailable = [
        row["standardized_field"]
        for row in mapping_rows
        if not row["mapping_status"].startswith("detected")
    ]
    mapping_payload = json.dumps(
        mapping_for_hash,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    source_url = (
        f"{downloader.BASE_URL}/{era_code}/data/seibi/facility/01{path.suffix.lower()}"
    )
    workbook_row = {
        "fiscal_year": fiscal_year,
        "era_code": era_code,
        "source_url": source_url,
        "source_url_status": "configured_downloader_pattern_not_network_revalidated",
        "filename": path.name,
        "repository_path": path.relative_to(REPO_ROOT).as_posix(),
        "file_status": "present",
        "file_nonempty": True,
        "sha256": sha256_file(path),
        "byte_size": int(stat.st_size),
        "retrieval_timestamp_utc": "unavailable",
        "retrieval_timestamp_status": "not_recorded_by_downloader",
        "file_mtime_utc": "unavailable_not_persisted",
        "repository_commit_timestamp": repository_commit_timestamp(path),
        "timestamp_basis": "git_last_commit_not_original_retrieval_time",
        "workbook_format": path.suffix.lower().lstrip("."),
        "parser_engine": engine,
        "workbook_sheet_count": len(sheet_names),
        "workbook_sheet_names_json": json.dumps(sheet_names, ensure_ascii=False),
        "parser_sheet_index_zero_based": 0,
        "parser_sheet_name": sheet_names[0],
        "raw_sheet_nonempty": True,
        "raw_sheet_rows": int(raw.shape[0]),
        "raw_sheet_columns": int(raw.shape[1]),
        "data_start_row_zero_based": data_start,
        "data_start_row_excel": data_start + 1,
        "candidate_data_rows": candidate_data_rows,
        "detected_standardized_fields": detected,
        "unavailable_standardized_fields": len(unavailable),
        "unavailable_field_names": ";".join(unavailable) if unavailable else "none",
        "schema_mapping_sha256": hashlib.sha256(mapping_payload).hexdigest(),
    }
    return workbook_row, mapping_rows


def validate_preflight_frames(
    workbooks: pd.DataFrame,
    mappings: pd.DataFrame,
    parser: ModuleType,
) -> None:
    """Validate complete inventory and per-year required schema detection."""
    expected_years = list(CANONICAL_FISCAL_YEARS)
    workbook_years = workbooks["fiscal_year"].astype(int).tolist()
    duplicate_workbook_years = sorted(
        year for year, count in Counter(workbook_years).items() if count > 1
    )
    missing_workbook_years = sorted(set(expected_years) - set(workbook_years))
    unexpected_workbook_years = sorted(set(workbook_years) - set(expected_years))
    if (
        len(workbooks) != EXPECTED_WORKBOOK_COUNT
        or workbook_years != expected_years
        or duplicate_workbook_years
    ):
        raise IngestionPreflightError(
            "Inspected workbook inventory is not canonical: "
            f"count={len(workbooks)}, expected={EXPECTED_WORKBOOK_COUNT}, "
            f"duplicates={duplicate_workbook_years or 'none'}, "
            f"missing={missing_workbook_years or 'none'}, "
            f"unexpected={unexpected_workbook_years or 'none'}"
        )

    if not workbooks["file_status"].eq("present").all():
        invalid = workbooks.loc[
            ~workbooks["file_status"].eq("present"),
            ["fiscal_year", "file_status"],
        ].to_dict("records")
        raise IngestionPreflightError(
            f"Inspected workbook inventory contains non-present entries: {invalid}"
        )
    if not workbooks["file_nonempty"].eq(True).all():
        years = workbooks.loc[
            ~workbooks["file_nonempty"].eq(True), "fiscal_year"
        ].astype(int).tolist()
        raise IngestionPreflightError(
            "Inspected workbook inventory contains empty files: "
            + ", ".join(f"FY{year}" for year in years)
        )
    if not workbooks["raw_sheet_nonempty"].eq(True).all() or (
        workbooks["candidate_data_rows"].astype(int) <= 0
    ).any():
        years = workbooks.loc[
            ~workbooks["raw_sheet_nonempty"].eq(True)
            | (workbooks["candidate_data_rows"].astype(int) <= 0),
            "fiscal_year",
        ].astype(int).tolist()
        raise IngestionPreflightError(
            "Inspected workbooks contain no candidate data rows: "
            + ", ".join(f"FY{year}" for year in years)
        )

    configured_fields = [field_name for field_name, _ in parser.COLUMN_DEFS]
    expected_mapping_count = EXPECTED_WORKBOOK_COUNT * len(configured_fields)
    duplicate_mappings = mappings.duplicated(
        ["fiscal_year", "standardized_field"], keep=False
    )
    if len(mappings) != expected_mapping_count or duplicate_mappings.any():
        duplicate_pairs = sorted(
            {
                (int(row.fiscal_year), str(row.standardized_field))
                for row in mappings.loc[duplicate_mappings].itertuples(index=False)
            }
        )
        raise IngestionPreflightError(
            "Schema map does not contain exactly one row per year/field: "
            f"rows={len(mappings)}, expected={expected_mapping_count}, "
            f"duplicates={duplicate_pairs or 'none'}"
        )

    expected_pairs = {
        (fiscal_year, field_name)
        for fiscal_year in CANONICAL_FISCAL_YEARS
        for field_name in configured_fields
    }
    actual_pairs = {
        (int(row.fiscal_year), str(row.standardized_field))
        for row in mappings.itertuples(index=False)
    }
    if actual_pairs != expected_pairs:
        missing_pairs = sorted(expected_pairs - actual_pairs)
        unexpected_pairs = sorted(actual_pairs - expected_pairs)
        raise IngestionPreflightError(
            "Schema map year/field coverage is not canonical: "
            f"missing={missing_pairs or 'none'}, "
            f"unexpected={unexpected_pairs or 'none'}"
        )

    expected_detected_pairs = {
        (fiscal_year, field_name)
        for fiscal_year in CANONICAL_FISCAL_YEARS
        for field_name in parser.expected_source_fields(fiscal_year)
    }
    required_for_year = mappings.apply(
        lambda row: (
            int(row["fiscal_year"]), str(row["standardized_field"])
        ) in expected_detected_pairs,
        axis=1,
    )
    missing_required = mappings[
        required_for_year
        & ~mappings["mapping_status"].str.startswith("detected")
    ]
    if not missing_required.empty:
        by_year = (
            missing_required.groupby("fiscal_year", sort=True)["standardized_field"]
            .apply(lambda values: ",".join(sorted(values)))
            .to_dict()
        )
        details = "; ".join(
            f"FY{int(year)}:[{fields}]" for year, fields in by_year.items()
        )
        raise IngestionPreflightError(
            "Required standardized-field detection failed: " + details
        )


def build_preflight_frames(
    downloader: ModuleType,
    parser: ModuleType,
    raw_dir: Path = RAW_DIR,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Run canonical inventory/schema preflight without writing artifacts."""
    inventory = preflight_workbook_inventory(downloader, parser, raw_dir=raw_dir)
    workbook_rows: list[dict[str, Any]] = []
    mapping_rows: list[dict[str, Any]] = []
    for fiscal_year, era_code, path in inventory:
        workbook_row, workbook_mappings = inspect_workbook(
            fiscal_year, path, era_code, downloader, parser
        )
        workbook_rows.append(workbook_row)
        mapping_rows.extend(workbook_mappings)

    workbooks = pd.DataFrame(workbook_rows)
    mappings = pd.DataFrame(mapping_rows)
    validate_preflight_frames(workbooks, mappings, parser)
    return workbooks, mappings


def write_report(workbooks: pd.DataFrame, mappings: pd.DataFrame) -> None:
    present = workbooks[workbooks["file_status"].eq("present")].copy()
    lines = [
        "# Raw Data Provenance",
        "",
        "This artifact records recoverable provenance for the raw MOE facility workbooks. "
        "URLs are reconstructed only from the checked-in downloader configuration; this "
        "stage does not claim that the URLs were revalidated at run time.",
        "",
        "- Original retrieval timestamp: **unavailable** because the downloader did not record it.",
        "- Filesystem modification time: **unavailable/not persisted** because checkout mtimes are volatile.",
        "- Repository timestamp: last Git commit time for each workbook, not the original retrieval time.",
        "- Parser sheet: first workbook sheet (`sheet_name=0`), matching the checked-in parser.",
        "- Header mappings: reproduced with the parser's first-match keyword search over rows 0-5.",
        "- Canonical preflight: **passed** for exactly one present, nonempty workbook per year.",
        f"- Canonical window: FY{CANONICAL_FISCAL_YEARS[0]}-FY{CANONICAL_FISCAL_YEARS[-1]}.",
        f"- Configured fiscal years: {len(workbooks):,}",
        f"- Present workbooks: {len(present):,}",
        f"- Total present bytes: {int(pd.to_numeric(present['byte_size']).sum()):,}",
        "",
        "## Workbook Inventory",
        "",
    ]
    display = present[
        [
            "fiscal_year",
            "era_code",
            "filename",
            "byte_size",
            "sha256",
            "repository_commit_timestamp",
            "parser_sheet_name",
            "raw_sheet_rows",
            "data_start_row_excel",
            "candidate_data_rows",
            "detected_standardized_fields",
            "unavailable_standardized_fields",
        ]
    ].copy()
    display["sha256"] = display["sha256"].str.slice(0, 12) + "..."
    lines.append(display.to_markdown(index=False))
    lines.extend(
        [
            "",
            "## Mapping Coverage",
            "",
            "The full field-by-field mapping, including matched Japanese header text and "
            "zero-based/Excel coordinates, is in `output/raw_workbook_schema_map.csv`.",
            "",
        ]
    )
    missing = mappings[~mappings["mapping_status"].str.startswith("detected")]
    if missing.empty:
        lines.append("All configured standardized fields were detected in every workbook.")
    else:
        coverage = (
            missing.groupby(["fiscal_year", "standardized_field"], as_index=False)
            .size()
            .drop(columns="size")
        )
        lines.append(
            "Optional later-schema fields not detected by the configured parser search "
            "(all required fields passed preflight):"
        )
        lines.append("")
        lines.append(coverage.to_markdown(index=False))
    lines.extend(
        [
            "",
            "## Audit Boundary",
            "",
            "SHA-256 and byte size establish the identity of the files currently in the "
            "repository. They do not establish the date of download, HTTP response headers, "
            "publisher-side version history, or legal custody before the files entered this "
            "workspace; those fields are unavailable unless a separate acquisition log exists.",
        ]
    )
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    downloader = load_analysis_module(
        "facility_downloader_config", "01_download_facility_data.py"
    )
    parser = load_analysis_module("facility_parser_config", "02_parse_facility_panel.py")

    # Complete all fail-hard checks before writing any provenance artifact.
    workbooks, mappings = build_preflight_frames(downloader, parser)
    print(
        "Canonical preflight passed: "
        f"{len(workbooks)} configured/present/nonempty workbooks, "
        f"FY{CANONICAL_FISCAL_YEARS[0]}-FY{CANONICAL_FISCAL_YEARS[-1]}"
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbooks.to_csv(WORKBOOK_MANIFEST_PATH, index=False)
    mappings.to_csv(SCHEMA_MAP_PATH, index=False)
    write_report(workbooks, mappings)

    present = workbooks[workbooks["file_status"].eq("present")]
    manifest_path = write_stage_manifest(
        "02b_build_raw_data_manifest",
        inputs=[
            "code/analysis/01_download_facility_data.py",
            "code/analysis/02_parse_facility_panel.py",
            "data/raw/facility_annual",
        ],
        outputs=[
            "output/raw_data_manifest.csv",
            "output/raw_workbook_schema_map.csv",
            "output/raw_data_provenance.md",
        ],
        metadata={
            "configured_fiscal_years": int(len(workbooks)),
            "canonical_fiscal_years": list(CANONICAL_FISCAL_YEARS),
            "present_workbooks": int(len(present)),
            "nonempty_workbooks": int(workbooks["file_nonempty"].sum()),
            "nonempty_first_sheets": int(workbooks["raw_sheet_nonempty"].sum()),
            "missing_workbooks": 0,
            "duplicate_workbook_years": [],
            "required_standardized_fields": list(
                parser.REQUIRED_STANDARDIZED_FIELDS
            ),
            "optional_standardized_fields": list(
                parser.OPTIONAL_STANDARDIZED_FIELDS
            ),
            "optional_field_first_fiscal_year": dict(
                parser.OPTIONAL_FIELD_FIRST_FISCAL_YEAR
            ),
            "candidate_data_rows_by_fiscal_year": {
                str(int(row.fiscal_year)): int(row.candidate_data_rows)
                for row in workbooks.itertuples(index=False)
            },
            "total_present_bytes": int(pd.to_numeric(present["byte_size"]).sum()),
            "retrieval_timestamp_available": False,
            "timestamp_used": "git_last_commit_timestamp",
            "source_url_basis": "checked-in_downloader_configuration",
            "parser_mapping_basis": "checked-in_parser_keyword_search",
            "downloader_sha256": sha256_file(
                ANALYSIS_DIR / "01_download_facility_data.py"
            ),
            "parser_sha256": sha256_file(ANALYSIS_DIR / "02_parse_facility_panel.py"),
            "workbook_sha256": {
                str(int(row.fiscal_year)): row.sha256
                for row in present.itertuples(index=False)
            },
        },
    )
    print(f"Workbooks recorded: {len(present):,}/{len(workbooks):,}")
    print(f"Workbook manifest: {WORKBOOK_MANIFEST_PATH}")
    print(f"Schema map: {SCHEMA_MAP_PATH}")
    print(f"Provenance report: {REPORT_PATH}")
    print(f"Stage manifest: {manifest_path}")


if __name__ == "__main__":
    main()
